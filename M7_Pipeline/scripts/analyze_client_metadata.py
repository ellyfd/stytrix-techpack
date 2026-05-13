"""analyze_client_metadata.py — 讀 PDF cover universe,輸出 per-客戶 metadata audit xlsx

讀 pdf_cover_universe.summary.json (extract_pdf_cover_universe.py 輸出),
per-client 出每客戶的 metadata fields + 各 field 的 unique values。

只取 coverage% >= MIN_COVERAGE_PCT (default 30%) 的 fields,過濾 long-tail noise。

輸出 xlsx 兩個 sheet:
1. client_field_universe:client × field 的 coverage matrix
                          + 每 (client, field) 對應的 unique value count
2. field_value_attributes: 每 (client, field) 的所有 unique values 列表

跑:python scripts/analyze_client_metadata.py [--min-coverage 30]
Output: outputs/platform/22_client_pdf_metadata_audit.xlsx
"""
from __future__ import annotations
import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SUMMARY_PATH = ROOT / "outputs" / "platform" / "pdf_cover_universe.summary.json"
OUT_XLSX = ROOT / "outputs" / "platform" / "22_client_pdf_metadata_audit.xlsx"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--min-coverage", type=float, default=30.0,
                    help="只列 coverage%% 達此門檻的 fields (default 30)")
    ap.add_argument("--top-n-clients", type=int, default=22,
                    help="top N clients by n_pdfs (default 22)")
    args = ap.parse_args()

    if not SUMMARY_PATH.exists():
        print(f"[FAIL] {SUMMARY_PATH} not found", file=sys.stderr)
        print(f"       先跑: python scripts/extract_pdf_cover_universe.py", file=sys.stderr)
        return 1

    summary = json.load(open(SUMMARY_PATH, encoding="utf-8"))
    print(f"Loaded summary: {len(summary)} clients")

    # Top N clients by n_pdfs
    top_clients = sorted(summary.keys(), key=lambda c: -summary[c]["n_pdfs"])[:args.top_n_clients]
    print(f"Top {len(top_clients)} clients: {top_clients}")

    # Filter to high-coverage keys
    print(f"Filtering to keys with coverage% >= {args.min_coverage}")
    client_fields = {}  # client → {field: stat}
    all_fields = set()
    for client in top_clients:
        keys = summary[client]["keys"]
        kept = {k: v for k, v in keys.items() if v["coverage_pct"] >= args.min_coverage}
        client_fields[client] = kept
        all_fields.update(kept.keys())
    all_fields = sorted(all_fields)
    print(f"Unique high-coverage fields across {len(top_clients)} clients: {len(all_fields)}")

    # Build xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[!] openpyxl not installed", file=sys.stderr)
        return 1

    wb = Workbook()

    # === Sheet 1: client × field coverage matrix ===
    ws1 = wb.active
    ws1.title = "client_field_universe"
    header_row = ["client", "n_pdfs"] + all_fields
    ws1.append(header_row)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")
    GRAY = PatternFill("solid", fgColor="EEEEEE")

    for client in top_clients:
        n = summary[client]["n_pdfs"]
        row = [client, n]
        for fld in all_fields:
            stat = client_fields[client].get(fld)
            if stat is None:
                row.append("")
            else:
                row.append(stat["coverage_pct"])
        ws1.append(row)

    # Color cells by coverage
    for row_idx in range(2, len(top_clients) + 2):
        for col_idx in range(3, len(header_row) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v == "" or v is None:
                cell.fill = GRAY
            elif isinstance(v, (int, float)):
                if v >= 70: cell.fill = GREEN
                elif v >= 50: cell.fill = YELLOW
                else: cell.fill = RED

    ws1.freeze_panes = "C2"

    # === Sheet 2: field × unique values per client ===
    ws2 = wb.create_sheet("field_value_attributes")
    ws2.append(["client", "field", "n_pdfs_with_field", "coverage_pct",
                "n_unique_values", "top_5_values_with_count", "all_unique_values_sample"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for client in top_clients:
        for fld, stat in sorted(client_fields[client].items(),
                                key=lambda kv: -kv[1]["coverage_pct"]):
            top5_str = " | ".join(
                f"{x['v'][:50]} ({x['count']})"
                for x in stat["top_5_values"]
            )
            all_uniq = " | ".join(stat["all_unique_values_sample"][:30])
            ws2.append([
                client,
                fld,
                stat["n_pdfs_with_key"],
                stat["coverage_pct"],
                stat["n_unique_values"],
                top5_str,
                all_uniq,
            ])
    ws2.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    print(f"\n✓ Wrote {OUT_XLSX}")
    print(f"  Sheets: client_field_universe / field_value_attributes")
    print(f"  Filter: coverage% >= {args.min_coverage}")

    # Print quick summary
    print(f"\n=== Summary by client (高 coverage fields 數量) ===")
    print(f"{'client':25} {'n_pdfs':>7} {'high_cov_fields':>16}  top 5")
    for client in top_clients:
        n = summary[client]["n_pdfs"]
        hi_fields = client_fields[client]
        nf = len(hi_fields)
        top5 = sorted(hi_fields.items(), key=lambda kv: -kv[1]["coverage_pct"])[:5]
        top5_str = " ".join(f"{k}({stat['coverage_pct']:.0f}%)" for k, stat in top5)
        print(f"{client:25} {n:>7} {nf:>16}  {top5_str}")

    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
