"""propose_canonical_xwalk.py — 22 客戶 PDF metadata 跟聚陽 M7 列管 canonical 完整對照

讀 pdf_cover_universe.summary.json,per-(client, field) 提議對到的 M7 canonical key + confidence。
輸出 xlsx 一個 sheet:`canonical_crosswalk` 含全部 high-coverage fields,user review 後填 final mapping。

跑:python scripts/propose_canonical_xwalk.py
Output: outputs/platform/canonical_crosswalk.xlsx
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SUMMARY = ROOT / "outputs" / "platform" / "pdf_cover_universe.summary.json"
OUT_XLSX = ROOT / "outputs" / "platform" / "canonical_crosswalk.xlsx"

MIN_COV = 30  # 只列 coverage >= 30% 的 fields

# Canonical M7 列管 keys (user 指定 only 8 個)
CANONICAL = [
    "客戶",            # client/brand
    "報價款號",        # 客戶款號
    "Program",         # 客戶 program(M7 內部)
    "Subgroup",        # 業務分類(M7 內部)
    "W/K",             # Knit/Woven
    "Item",            # PullOn / Leggings 等
    "Season",          # 季
    "PRODUCT_CATEGORY", # gender (Women/Men/Girl/Boy/Baby)
    "(no map)",        # 不對應 8 個 canonical 的全標 no map
]

# Auto suggest based on field name keyword + value pattern
# 限制 only 8 個 canonical:客戶/報價款號/Program/Subgroup/W/K/Item/Season/PRODUCT_CATEGORY
def auto_suggest(client: str, field: str, top_values: list, n_unique: int) -> tuple[str, str, str]:
    """Returns (suggested_canonical, confidence, notes)
    confidence: HIGH / MEDIUM / LOW
    """
    fl = field.lower().strip(":：# ")
    top_v_str = " ".join(v["v"].lower() for v in top_values[:5])
    sample_vals = " | ".join(v["v"][:30] for v in top_values[:3])

    # === 報價款號 (style number) ===
    # 注意:CATO VPN 跟 DICKS Sample Season top value 看起來像 noise,降為 LOW
    if fl in ("vpn", "sample season"):
        # CATO VPN top value = "Season: 2025 Fall Apparel" (detector 抓錯)
        # DICKS Sample Season top value = "Proto - 200797 -" (sample id, not season)
        return "(no map)", "LOW", f"detector 抓錯結構,top value 不像 style/season"
    if any(kw in fl for kw in [
        "style #", "style#", "style no", "style code", "style number",
        "デザイン名", "款号", "item trading code",
    ]):
        # 加 value sanity check:含字母+數字混合,長度 5-20 才算 style number
        valid_style = any(re.match(r"^[A-Z0-9_\-]{5,25}$", v["v"]) for v in top_values[:3])
        return ("報價款號", "HIGH" if valid_style else "MEDIUM", "客戶款號")
    if "tech pack" in fl and "number" in fl:
        return "報價款號", "HIGH", ""
    if fl == "style description" and client == "CATO":
        return "報價款號", "MEDIUM", "CATO 用 Style Description (檢查值不是純描述)"
    if fl == "bom number":
        # 聚陽 HEADER_SN ≠ 報價款號,降為 (no map)
        return "(no map)", "LOW", "Centric 8 BOM number,M7 沒對應(HEADER_SN 是聚陽內部)"

    # === 客戶 (brand) ===
    if any(kw in fl for kw in [
        "brand/", "brand division", "brand category", "brand-",
        "ブランド", "品牌", "company name", "企業",
    ]):
        return "客戶", "HIGH", ""
    if fl == "brand":
        return "客戶", "HIGH", ""

    # === Season ===
    if fl == "season" or fl == "シーズン":
        # value sanity:含 SS/FW/Fall/Spring/Summer/Holiday/年份 才算
        valid = any(re.search(r"(SS|FW|AW|Fall|Spring|Summer|Holiday|Winter|20\d\d)", v["v"], re.I)
                    for v in top_values[:3])
        return ("Season", "HIGH" if valid else "LOW", "")
    if "design season" in fl:
        return "Season", "HIGH", ""

    # === Item ===
    if any(kw in fl for kw in ["品类", "item trading code"]):
        return "Item", "HIGH", ""
    if fl == "アイテム":
        return "Item", "HIGH", ""
    if fl == "size category":
        return "Item", "MEDIUM", "CATO 用 Size Category 區分款型(Missy/Woman 等)"

    # === PRODUCT_CATEGORY (gender) ===
    if fl in ("group", "gender", "department"):
        # 看 value 確認
        if any(g in top_v_str for g in ["women", "men", "girl", "boy", "baby", "kid", "ladies"]):
            return "PRODUCT_CATEGORY", "HIGH", "value 含 gender keyword"
        return ("(no map)", "LOW", "field 名像但 value 不含 gender")
    if "product team" in fl:
        if any(g in top_v_str for g in ["women", "men", "girl", "boy"]):
            return "PRODUCT_CATEGORY", "MEDIUM", "客戶 team 含 gender"
        return ("(no map)", "LOW", "")

    # === W/K (fabric type) ===
    if any(kw in fl for kw in [
        "body fabric", "main fabric", "main material", "shell fabric", "fabric content",
    ]):
        return "W/K", "MEDIUM", "面料 spec → 推 Knit/Woven(用 keyword 比對)"
    if fl == "design type":
        if any(kw in top_v_str for kw in ["knit", "woven"]):
            return "W/K", "HIGH", ""
        return ("(no map)", "LOW", "")
    if fl == "category" and any(kw in top_v_str for kw in ["knit", "woven", "circular"]):
        return "W/K", "MEDIUM", "客戶 category 含 KNIT/WOVEN/CIRCULAR(可推 W/K)"

    # === Program / Subgroup === (M7 內部分類,PDF 通常沒)
    # No auto suggest for these (極少客戶 PDF 對得上)

    # === 全部不對應 8 canonical 的標 (no map) ===
    return "(no map)", "LOW", f"不對應 8 canonical / 待人工確認 / sample: {sample_vals[:80]}"


def main():
    sm = json.load(open(SUMMARY, encoding="utf-8"))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl required: pip install openpyxl")
        return 1

    wb = Workbook()
    ws = wb.active
    ws.title = "canonical_crosswalk"
    ws.append([
        "client", "n_pdfs", "field_name", "coverage_pct",
        "n_unique_values", "top_5_values_with_count", "all_unique_values_sample",
        "suggested_canonical", "confidence", "auto_notes",
        "(IE 部門填) final_canonical", "(IE 部門填) notes",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    HIGH = PatternFill("solid", fgColor="C6EFCE")
    MED = PatternFill("solid", fgColor="FFEB9C")
    LOW = PatternFill("solid", fgColor="FFC7CE")
    NOMAP = PatternFill("solid", fgColor="EEEEEE")

    n_rows = 0
    by_client_n = {}
    by_canonical_n = {}
    by_confidence_n = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}

    # Sort: 22 clients by n_pdfs desc, within client by coverage desc
    clients = sorted(sm.keys(), key=lambda c: -sm[c]["n_pdfs"])[:22]
    for client in clients:
        info = sm[client]
        n_pdfs = info["n_pdfs"]
        keys = info["keys"]
        high = [(k, v) for k, v in keys.items() if v["coverage_pct"] >= MIN_COV]
        if not high:
            continue
        by_client_n[client] = len(high)
        # Sort within client by coverage desc
        for k, v in sorted(high, key=lambda kv: -kv[1]["coverage_pct"]):
            top5_str = " | ".join(f"{x['v'][:50]}({x['count']})" for x in v["top_5_values"])
            all_uniq_str = " | ".join(v["all_unique_values_sample"][:30])
            canon, conf, notes = auto_suggest(client, k, v["top_5_values"], v["n_unique_values"])
            ws.append([
                client, n_pdfs, k, v["coverage_pct"],
                v["n_unique_values"], top5_str, all_uniq_str,
                canon, conf, notes, "", "",
            ])
            n_rows += 1
            by_canonical_n[canon] = by_canonical_n.get(canon, 0) + 1
            by_confidence_n[conf] = by_confidence_n.get(conf, 0) + 1

            # Color by confidence
            row_idx = n_rows + 1
            cell = ws.cell(row=row_idx, column=9)  # confidence col
            if conf == "HIGH":
                cell.fill = HIGH
            elif conf == "MEDIUM":
                cell.fill = MED
            elif conf == "LOW":
                cell.fill = LOW

            # Color suggested_canonical
            cell2 = ws.cell(row=row_idx, column=8)
            if canon == "(no map)" or canon == "(待人工確認)":
                cell2.fill = NOMAP

    ws.freeze_panes = "A2"
    # Adjust column widths
    widths = [25, 8, 30, 10, 8, 80, 80, 25, 12, 50, 25, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    wb.save(OUT_XLSX)

    # Console summary
    print(f"\n=== Canonical Crosswalk Complete ===")
    print(f"  Total rows: {n_rows}")
    print(f"  XLSX: {OUT_XLSX}")
    print()
    print(f"=== By client (高 coverage fields >= {MIN_COV}%) ===")
    for c in sorted(by_client_n.keys(), key=lambda x: -by_client_n[x]):
        print(f"  {c:30} {by_client_n[c]:>3} fields")
    print()
    print(f"=== By suggested canonical ===")
    for c, n in sorted(by_canonical_n.items(), key=lambda x: -x[1]):
        print(f"  {c:30} {n:>3} field-rows")
    print()
    print(f"=== By confidence ===")
    for c, n in by_confidence_n.items():
        print(f"  {c:10} {n:>4}")


if __name__ == "__main__":
    main()
