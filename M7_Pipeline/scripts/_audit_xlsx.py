"""Audit xlsx_facets.jsonl: 看 4 件 ok 但 0 callouts 的是什麼 xlsx."""
import json
from pathlib import Path

XLSX_JSONL = Path(r"C:\Users\ellycheng.DOMAIN\Desktop\StyTrix TechPack\Construction\Source-Data\M7_Pipeline\outputs\extract\xlsx_facets.jsonl")
OUT = Path(r"C:\Users\ellycheng.DOMAIN\Desktop\StyTrix TechPack\Construction\Source-Data\M7_Pipeline\xlsx_audit.txt")

lines = []
def emit(s=""):
    print(s)
    lines.append(s)

emit("=" * 70)
emit("  xlsx 4 件 ok audit")
emit("=" * 70)

n_ok = 0
with open(XLSX_JSONL, encoding="utf-8") as f:
    for line in f:
        d = json.loads(line)
        if d.get("_status") != "ok":
            continue
        n_ok += 1
        emit(f"\n[{n_ok}] EIDH={d.get('eidh')}  client={d.get('client_code')}")
        emit(f"    xlsx_files: {d.get('xlsx_files')}")
        emit(f"    iso_callouts: {len(d.get('iso_callouts', []))}")
        emit(f"    mcs: {len(d.get('mcs', []))}")

# 接下來自己 open 第一個 xlsx 看 sheet structure
emit(f"\n=== 抽第一個 xlsx 看 sheet structure ===")
import openpyxl
with open(XLSX_JSONL, encoding="utf-8") as f:
    for line in f:
        d = json.loads(line)
        if d.get("_status") != "ok": continue
        xlsx_files = d.get("xlsx_files", [])
        if not xlsx_files: continue
        eidh = d.get("eidh")
        # find folder
        from pathlib import Path
        tp = Path(r"C:\Users\ellycheng.DOMAIN\Desktop\StyTrix TechPack\Construction\Source-Data\M7_Pipeline\tp_samples_v2")
        for folder in tp.iterdir():
            if not folder.is_dir(): continue
            if folder.name.startswith(eidh + "_"):
                xlsx_path = folder / xlsx_files[0]
                if not xlsx_path.exists(): continue
                emit(f"\n  抽: {xlsx_path.name}")
                try:
                    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
                    emit(f"    sheets: {wb.sheetnames}")
                    for sn in wb.sheetnames[:3]:
                        ws = wb[sn]
                        emit(f"    [{sn}] (max_row≈{ws.max_row})")
                        rows = list(ws.iter_rows(max_row=20, values_only=True))
                        for i, row in enumerate(rows[:12]):
                            cells = [str(c) if c is not None else "" for c in row[:6]]
                            emit(f"      row{i}: {cells}")
                except Exception as e:
                    emit(f"    [err] {e}")
                break
        break  # 只看第一個

OUT.write_text("\n".join(lines), encoding="utf-8")
emit(f"\n[output] {OUT}")
