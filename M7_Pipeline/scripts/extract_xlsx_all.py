"""extract_xlsx_all.py — Unified xlsx Techpack extractor.

對 tp_samples_v2/<EIDH>/*.xlsx 抽:
  - translation sheet → ISO callout (zone × description × ISO code) 結構化 fact
  - measurement chart sheet → mc dict (body_type/sizes/poms) — future expansion
  - junk sheet → skip

xlsx 結構比 PDF 簡單 (一份 xlsx 多數只一種 sheet 內容), 不需要 per-row 分類 dispatch,
只在 sheet 層級 classify 然後 parse 該 sheet 全表。

來源 邏輯:抽自 extract_xlsx_callout.py, 加 measurement chart sheet 偵測。

用法:
  python scripts/extract_xlsx_all.py [--limit N] [--workers N] [--reset]
"""
from __future__ import annotations
import argparse
import json
import re
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from collections import Counter

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent
TP_DIR = ROOT / "tp_samples_v2"
MANIFEST_PATH = ROOT / "m7_organized_v2" / "_fetch_manifest.csv"
OUT_DIR = ROOT / "outputs" / "extract"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_JSONL = OUT_DIR / "xlsx_facets.jsonl"

ISO_RE = re.compile(r"\b(301|401|406|503|504|512|514|515|516|601|602|605|607)\b")
SKIP_SHEETS = ["instruction", "instructions", "cover", "summary"]
TRANSLATION_HINTS = ["針法", "做工", "iso standard", "construction", "construction"]
MC_HINTS = ["measurement", "尺寸", "規格", "tolerance", "pom"]


# Folder name parsing (與 pdf_all/pptx_all 共用邏輯)
KNOWN_CLIENT_TOKENS = [
    "DICKS_SPORTING_GOODS", "ABERCROMBIE_AND_FITCH",
    "OLD_NAVY", "GAP_OUTLET", "BANANA_REPUBLIC", "WAL-MART-CA",
    "GAP", "DICKS", "ATHLETA", "UNDER_ARMOUR", "KOHLS", "A_AND_F", "GU", "BEYOND_YOGA",
    "HIGH_LIFE_LLC", "WAL-MART", "QUINCE", "HALARA", "NET",
    "JOE_FRESH", "BRFS", "SANMAR", "DISTANCE", "ZARA",
    "ASICS-EU", "TARGET", "LEVIS", "CATO", "SMART_CLOTHING",
]
CLIENT_RAW_TO_CODE = {
    "OLD NAVY": "ONY", "TARGET": "TGT", "GAP": "GAP", "GAP OUTLET": "GAP",
    "DICKS SPORTING GOODS": "DKS", "DICKS": "DKS", "ATHLETA": "ATH",
    "UNDER ARMOUR": "UA", "KOHLS": "KOH", "A AND F": "ANF", "A & F": "ANF",
    "GU": "GU", "BEYOND YOGA": "BY", "HIGH LIFE LLC": "HLF", "WAL-MART": "WMT",
    "WAL-MART-CA": "WMT", "QUINCE": "QCE", "HALARA": "HLA", "NET": "NET",
    "JOE FRESH": "JF", "BANANA REPUBLIC": "BR", "BRFS": "BR", "SANMAR": "SAN",
    "DISTANCE": "DST", "ZARA": "ZAR", "ASICS-EU": "ASICS", "LEVIS": "LEV",
    "CATO": "CATO", "SMART CLOTHING": "SMC",
}


def _load_manifest_lookup() -> dict:
    import csv
    lookup = {}
    if not MANIFEST_PATH.exists():
        return lookup
    with open(MANIFEST_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            eidh = (row.get("Eidh") or "").strip()
            if eidh:
                lookup[eidh] = {
                    "客戶": (row.get("客戶") or "").strip(),
                    "報價款號": (row.get("報價款號") or "").strip(),
                    "Item": (row.get("Item") or "").strip(),
                    "HEADER_SN": (row.get("HEADER_SN") or "").strip(),
                }
    return lookup


_MANIFEST_LOOKUP = {}


def _init_worker(manifest_lookup):
    global _MANIFEST_LOOKUP
    _MANIFEST_LOOKUP = manifest_lookup


def _parse_folder_name(folder_name: str, manifest_lookup: dict = None) -> dict:
    lookup = manifest_lookup if manifest_lookup is not None else _MANIFEST_LOOKUP
    parts = folder_name.split("_", 1)
    eidh = parts[0] if parts else None
    design_suffix = parts[1] if len(parts) > 1 else ""
    info = lookup.get(eidh, {})
    client_raw = info.get("客戶", "")
    design_id = info.get("報價款號") or design_suffix
    client_code = CLIENT_RAW_TO_CODE.get(client_raw.upper().strip(), client_raw[:6].upper() if client_raw else "UNKNOWN")
    return {
        "eidh": eidh,
        "hsn": info.get("HEADER_SN", ""),
        "client_raw": client_raw,
        "client_code": client_code,
        "design_id": design_id,
        "item": info.get("Item", ""),
    }


def _parse_gu_is_sheet(sheet_name: str, rows: list[list]) -> list[dict]:
    """GU 日文式樣書 IS_JP / IS_US 工作表 → 抽 POM rows.

    2026-05-12 加: GU XLSX 主源是 IS (Inspection Sheet) 式樣書, 22 sheet 結構.
    跟標準 Centric 8 POM table 不同 (日英混合, 縱列部位 / 橫列 size+grading).

    策略 (heuristic, 因 GU 多 sheet 變體):
      1. Scan 找含 size header (XS/S/M/L/XL 或 サイズ) 的 row → header_row_idx
      2. Header 之下 row 是 POM 行 (part_name + grading values)
      3. Part name 可能是日英混合 (e.g. "着丈/Body Length")

    Returns: list of POM dicts {POM_Code, POM_Name, sizes, tolerance, _raw_part}
    """
    if not rows:
        return []

    poms = []
    rows = list(rows)

    # === Step 1: 找 size header row ===
    SIZE_TOKENS = {"XXS", "XS", "S", "M", "L", "XL", "XXL", "3XL", "4XL", "5XL",
                   "サイズ", "size", "Size", "SIZE"}
    NUM_SIZE_RE = re.compile(r"^\d{1,3}$")  # 數字 size (e.g. 90/95/100/M/L)

    header_idx = -1
    header_cols = []  # [(col_idx, size_label)]
    for i, row in enumerate(rows[:60]):  # GU header 通常前 60 row
        size_matches = []
        for ci, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            if s in SIZE_TOKENS or NUM_SIZE_RE.match(s):
                size_matches.append((ci, s))
        # 至少 3 個 size token 同 row 才算 header
        if len(size_matches) >= 3:
            header_idx = i
            header_cols = size_matches
            break

    if header_idx < 0:
        return []  # 沒找到 size header

    # === Step 2: header 下面是 POM rows ===
    # part name 通常在 row 第 1-2 cell
    NUMBER_RE = re.compile(r"^-?\d+(?:\.\d+)?(?:\s*[/／]\s*\d+(?:\.\d+)?)?\s*[\"”]?$")
    # ex: "60", "60.5", "1 1/2", "1/2"

    for ri in range(header_idx + 1, min(header_idx + 60, len(rows))):
        row = rows[ri]
        if not row: continue
        # First non-empty cell = part name
        part_name = ""
        first_data_col = -1
        for ci, cell in enumerate(row):
            if cell is None: continue
            s = str(cell).strip()
            if not s: continue
            if NUMBER_RE.match(s):
                if first_data_col < 0:
                    first_data_col = ci
                break  # found first numeric cell
            else:
                # Treat as part name
                if not part_name:
                    part_name = s[:80]

        if not part_name:
            continue  # not a POM row

        # Map header_cols to values in this row
        sizes_dict = {}
        for col_idx, size_label in header_cols:
            if col_idx < len(row) and row[col_idx] is not None:
                v = str(row[col_idx]).strip()
                if v and NUMBER_RE.match(v):
                    sizes_dict[size_label] = v

        if sizes_dict and len(sizes_dict) >= 2:  # 至少 2 個 size 有值才算
            poms.append({
                "POM_Code": f"GU{ri - header_idx:03d}",  # 自動編碼
                "POM_Name": part_name,
                "sizes": sizes_dict,
                "tolerance": {},
                "_raw_part": part_name,
            })

    return poms


def _parse_metadata_sheet(sheet_name: str, rows: list[list], client_code: str) -> dict:
    """從 XLSX 抽 metadata (WMT/SAN/QCE/NET 4 brand cover layouts).

    2026-05-12 加: WMT/SAN/QCE/NET 為 XLSX 主源, PDF 無 metadata. 從 XLSX cover sheet 抽
    style_number / customer / brand / season / department / buyer / size_range / description 等.

    Per-brand detection:
      WMT (Walmart):  "做工翻譯" / "參考照片" sheet, row 8-12 含 VENDOR/STYLE #/SEASON NAME/BRAND/DEPARTMENT/BUYER
      SAN (Sanmar):   "P2 comments" / Sample Evaluation Form, row 3-7 含 Style # / Ref. Sample Log# / Descriptions / Size
      QCE (Quince):   per-style sheet, row 1-5 含 QUINCE / Product Type / Gender / Style Number / Description
      NET (Net):      "尺寸表", row 1-5 含「主富服裝」/ STYLE NO:
    """
    meta = {}
    # Scan first 20 rows
    rows = list(rows)[:20]

    def get_cell(row_idx, col_idx, default=""):
        if 0 <= row_idx < len(rows):
            r = rows[row_idx]
            if 0 <= col_idx < len(r):
                v = r[col_idx]
                return str(v).strip() if v is not None else default
        return default

    def find_value_after_key(key_pattern: str, max_distance: int = 8) -> str:
        """從 rows 找 cell 內含 key 後, 同 row 或下 row 的 value.

        2026-05-12 改: 跳過「看起來是 key」的 cell (結尾 ':' 或全 UPPERCASE 且短),
                     避免把 next key 當 value (WMT layout 左右並列 key 容易踩雷).
        """
        import re
        key_re = re.compile(key_pattern, re.IGNORECASE)

        def looks_like_key(s: str) -> bool:
            s = s.strip()
            if not s:
                return True  # empty: skip
            if s.endswith(":") or s.endswith(":"):
                return True
            # 短全大寫詞 (e.g. "VENDOR" "BRAND") 也算 key-ish
            if len(s) < 25 and s.isupper() and " " not in s.strip(":"):
                return True
            return False

        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row):
                if cell is None:
                    continue
                s = str(cell).strip()
                if key_re.search(s):
                    # 看同 row 下 max_distance 個 cell, 跳過 key-like cells
                    for ci2 in range(ci + 1, min(ci + 1 + max_distance, len(row))):
                        v = row[ci2]
                        if v is None:
                            continue
                        vs = str(v).strip()
                        if not vs or vs == "0":
                            continue
                        if looks_like_key(vs):
                            continue  # skip next key
                        return vs
                    # 或下一 row 同 column (WMT/SAN 多 row layout)
                    if ri + 1 < len(rows):
                        next_row = rows[ri + 1]
                        for ci2 in range(ci, min(ci + max_distance, len(next_row))):
                            if ci2 < len(next_row) and next_row[ci2] is not None:
                                vs = str(next_row[ci2]).strip()
                                if vs and vs != "0" and not looks_like_key(vs):
                                    return vs
        return ""

    # === Per-brand 偵測 + 抽取 ===
    blob = " ".join(
        " ".join(str(c) for c in row if c is not None)
        for row in rows
    )
    blob_upper = blob.upper()

    if client_code == "WMT" or "做工翻譯" in sheet_name or ("VENDOR:" in blob_upper and "BUYER:" in blob_upper):
        # WMT layout
        if "VENDOR:" in blob_upper:
            meta["vendor"] = find_value_after_key(r"VENDOR:")[:50]
            meta["season_name"] = find_value_after_key(r"SEASON NAME:")[:50]
            meta["brand"] = find_value_after_key(r"BRAND:")[:30]
            meta["department"] = find_value_after_key(r"DEPARTMENT:")[:30]
            meta["buyer"] = find_value_after_key(r"BUYER:")[:50]
            meta["style_number"] = find_value_after_key(r"STYLE\s*#:")[:50]
            meta["prod_spec"] = find_value_after_key(r"PROD SPEC#:")[:50]
            meta["size_range"] = find_value_after_key(r"SIZE RANGE")[:50]
            meta["country"] = find_value_after_key(r"COUNTRY:")[:30]
            meta["description"] = find_value_after_key(r"DESCRIPTION:")[:120]
            # WMT 樣式: style code 通常在 row 3 中段 (e.g. "GRY5 8826LD_PD")
            style_code_cell = get_cell(2, 3) or get_cell(2, 4)
            if style_code_cell and len(style_code_cell) > 3:
                meta["design_number"] = style_code_cell[:50]

    if client_code == "SAN" or "Sample Evaluation" in blob or "Ref. Sample Log" in blob:
        # SAN layout
        if "STYLE #" in blob_upper.replace(" ", "") or "Style #" in blob:
            meta["style_number"] = find_value_after_key(r"Style\s*#")[:50]
            meta["design_number"] = meta.get("style_number", "")
            meta["ref_sample_log"] = find_value_after_key(r"Ref\.?\s*Sample\s*Log")[:50]
            meta["description"] = find_value_after_key(r"Descriptions?")[:120]
            meta["division"] = find_value_after_key(r"Division")[:30]
            meta["gender"] = find_value_after_key(r"Gender")[:30]
            meta["vendor"] = find_value_after_key(r"Vendor")[:50]
            meta["size_range"] = find_value_after_key(r"Size")[:30]

    if client_code == "QCE" or "QUINCE" in blob_upper:
        # QCE layout
        if "Style Number" in blob or "QUINCE" in blob_upper:
            meta["style_number"] = find_value_after_key(r"Style Number")[:50]
            meta["design_number"] = meta.get("style_number", "")
            meta["product_type"] = find_value_after_key(r"Product Type")[:30]
            meta["gender"] = find_value_after_key(r"Gender")[:30]
            meta["description"] = find_value_after_key(r"Description")[:120]
            meta["size_range"] = find_value_after_key(r"Size Range")[:30]
            meta["sample_size"] = find_value_after_key(r"Sample Size")[:30]
            meta["status"] = find_value_after_key(r"Status")[:30]
            meta["customer"] = "QUINCE"

    if client_code == "NET" or "主富服裝" in blob:
        # NET layout
        if "STYLE NO" in blob_upper or "主富服裝" in blob:
            meta["style_number"] = find_value_after_key(r"STYLE NO:")[:50]
            meta["design_number"] = meta.get("style_number", "")
            meta["customer"] = "主富服裝"

    return meta


def _classify_sheet(sheet_name: str, sample_rows: list[list]) -> str:
    """Return 'translation' / 'measurement' / 'junk'."""
    name_lower = sheet_name.lower()
    if any(s in name_lower for s in SKIP_SHEETS):
        return "junk"

    # Scan a few rows for hints
    blob = " ".join(
        " ".join(str(c) for c in row if c is not None)
        for row in sample_rows[:15]
    ).lower()

    if any(h in blob for h in TRANSLATION_HINTS) or any(h in name_lower for h in ["針", "做工", "construction"]):
        # 也要看有沒有 ISO code 出現過
        if ISO_RE.search(blob) or "iso" in blob:
            return "translation"
    # MC 偵測 (2026-05-12, 平衡版):
    # 之前太寬 (工作表1/P2 comments 誤判) → 收緊 → 太嚴格錯放真 MC
    # 平衡: name 級匹配 OR 內容含 POM table 信號 (size+tol)
    MC_NAME_KEYWORDS = ["基碼尺寸表", "全碼尺寸表", "尺寸表", "尺碼表", "尺碼", "尺寸",
                       "measurement chart", "measurement sheet", "measurements",
                       "fit mmt sheet", "mc sheet", "pom chart", "pom sheet",
                       "fit comments", "fit review", "size", "spec sheet", "spec",
                       "graded", "grade rule"]
    name_is_mc = any(h in name_lower for h in MC_NAME_KEYWORDS)
    if name_is_mc:
        return "measurement"
    # 內容偵測: 同 row 含 size token + tol + 數字 → 視為 MC
    # 細化: 需要 POM-ish row (有編號 + measurement-like data) 而不是隨意 "POM" 字
    has_size_label = any(s in blob for s in
                         ["xxs", "\nxs", " xs ", "\ns ", " s\n", "size:", "尺寸",
                          "2t", "3t", "4t", "yxs", "ysm"])
    has_tol = ("tol" in blob or "tolerance" in blob or "+/-" in blob)
    has_pom_or_meas = ("pom" in blob or "measurement" in blob or "尺寸" in blob or
                       "點" in blob or "circumference" in blob or "length" in blob or
                       "width" in blob)
    if has_size_label and has_tol and has_pom_or_meas:
        return "measurement"
    return "junk"


# Keyword → ISO mapping (from Data lookup tables found in real xlsx files)
# Used to infer ISO from English description text when no inline number
DESC_TO_ISO = [
    # (regex pattern, primary ISO)
    (r"\bLOCKSTITCH\b|\bSINGLE NEEDLE LOCK\b|\bTOPSTITCH\b|\bSINGLE NEEDLE\b", "301"),
    (r"\bZIG\s*ZAG\b.*LOCKSTITCH", "304"),
    (r"\bCHAIN\s*STITCH\b|\b2\s*THREAD CHAIN\b", "401"),
    (r"\bBOTTOM\s*COVERSTITCH\b|\b2\s*NEEDLE BOTTOM\b", "406"),
    (r"\b3\s*NEEDLE BOTTOM COVER", "407"),
    (r"\bOVEREDGE\b|\b2\s*THREAD OVEREDGE\b", "503"),
    (r"\b3\s*THREAD OVEREDGE\b|\bOVERLOCK\b|\bSERGE\b", "504"),
    (r"\b4\s*THREAD OVEREDGE\b|\b2\s*NEEDLE 4\s*THREAD\b", "514"),
    (r"\b4\s*THREAD SAFETY\b", "515"),
    (r"\b5\s*THREAD SAFETY\b", "516"),
    (r"\bCOVERSTITCH\b(?!.*BOTTOM)|\b2\s*NEEDLE COVER\b", "602"),
    (r"\b3\s*NEEDLE COVERSTITCH\b|\b3\s*NEEDLE 5\s*THREAD\b", "605"),
    (r"\bFLATLOCK\b|\bFLAT\s*SEAM\b|\b606\b", "606"),
    (r"\b607\b|\b4\s*NEEDLE 6\s*THREAD\b", "607"),
    (r"\bBLIND\s*STITCH\b|\bBLIND\s*HEM\b", "103"),
    (r"\bBAR\s*TACK\b", "304"),
    # Generic operations
    (r"\bBINDING\b|\bENCASED\b", "406"),  # binding 通常 cover stitch
    (r"\bHEMMING\b|\bHEM\b\s*(STITCH|FINISH)", "406"),
]

import re as _re_xlsx
_DESC_TO_ISO_COMPILED = [(_re_xlsx.compile(pat, _re_xlsx.IGNORECASE), iso) for pat, iso in DESC_TO_ISO]


def _infer_iso_from_description(desc: str) -> list[str]:
    """從描述文字推 ISO code (可能多個)."""
    if not desc:
        return []
    isos = []
    seen = set()
    for pat, iso in _DESC_TO_ISO_COMPILED:
        if pat.search(desc) and iso not in seen:
            isos.append(iso)
            seen.add(iso)
    return isos


def _find_header_row(sheet_rows: list[list]) -> int:
    """找真實表頭 row index (要求 ≥2 個 header keyword 出現在 *不同 cell*).

    避免誤觸 cover 那行的單一 'DESCRIPTION:' 標籤 — 該 row 只有 1 cell 含 keyword.
    真實 header 通常 ≥2 cell 各含一個 column name (Location | Description / Construction 等).
    """
    HEADER_KW = ("LOCATION", "DESCRIPTION", "CONSTRUCTION", "POINT OF MEASURE",
                 "STITCH TYPE", "PART NAME", "ZONE", "PART", "工法", "做工", "縫法",
                 "POM", "MEASUREMENT", "STITCH")
    BLACKLIST_LABEL = ("DESCRIPTION:", "STYLE DESCRIPTION", "CONSTRUCTION:")
    for i, row in enumerate(sheet_rows):
        if not row or i > 30:
            break
        if not any(c for c in row):
            continue
        cells_upper = [str(c).upper() if c is not None else "" for c in row]
        # 排除 cover/label row (cells ending with ":" are typically labels not headers)
        if any(blk in c for c in cells_upper for blk in BLACKLIST_LABEL):
            continue
        # 計算 keyword cell 數 (≥2 個 *不同* cell 含 header keyword)
        cells_with_kw = sum(1 for c in cells_upper if c and any(kw in c for kw in HEADER_KW))
        if cells_with_kw >= 2:
            return i
        # 也接受 single cell 但有"Location"+"Description"二合一文字
        for c in cells_upper:
            if "LOCATION" in c and "DESCRIPTION" in c:
                return i
    return -1


def _parse_translation_sheet(sheet_rows: list[list]) -> list[dict]:
    """從 translation sheet 抽 ISO callout facts.

    新策略 (2026-05-12):
    1. 偵測 header row (含 'Location' / 'Description' / 'Construction')
    2. 識別 zone column (Location) 和 description column (Description / Construction)
    3. 對每 data row 抽 zone + description, 用 keyword 推 ISO
    4. 若 description 內有 inline ISO 數字也保留
    """
    facts = []
    header_row_idx = _find_header_row(sheet_rows)
    if header_row_idx < 0:
        # Fallback: 原舊邏輯 (跳過前 12 行找 ISO inline)
        for i, row in enumerate(sheet_rows):
            if i < 12 or not row or all(c is None for c in row):
                continue
            cells = [str(c) if c is not None else "" for c in row]
            text_blob = " ".join(cells).upper()
            iso_match = ISO_RE.search(text_blob)
            if not iso_match:
                continue
            nonempty = [c for c in cells if c.strip()]
            if len(nonempty) < 2:
                continue
            facts.append({
                "zone_raw": nonempty[0][:80],
                "description": nonempty[1][:200],
                "iso": iso_match.group(1),
                "row_index": i,
            })
        return facts

    # 識別 zone / description 欄位 index
    header = [str(c) if c is not None else "" for c in sheet_rows[header_row_idx]]
    zone_col = -1
    desc_col = -1
    for ci, h in enumerate(header):
        h_upper = h.upper()
        if zone_col < 0 and ("LOCATION" in h_upper or "POSITION" in h_upper or "ZONE" in h_upper or "PART" in h_upper or "工法" in h or "部位" in h):
            zone_col = ci
        if desc_col < 0 and ("DESCRIPTION" in h_upper or "CONSTRUCTION" in h_upper or "做工" in h or "縫法" in h or "STITCH" in h_upper):
            desc_col = ci

    # 萬一沒找到 explicit columns, fallback to first 2 non-empty
    if zone_col < 0:
        zone_col = 0
    if desc_col < 0:
        # 找下一個非 zone_col 的非空 cell column
        for ci in range(len(header)):
            if ci != zone_col and (header[ci] or any(r[ci] for r in sheet_rows[header_row_idx+1:header_row_idx+5] if r and len(r) > ci)):
                desc_col = ci
                break

    # 抽 data rows
    for i in range(header_row_idx + 1, len(sheet_rows)):
        row = sheet_rows[i]
        if not row or all(c is None for c in row):
            continue
        if zone_col >= len(row) or desc_col >= len(row):
            continue
        zone = str(row[zone_col] or "").strip()
        desc = str(row[desc_col] or "").strip()
        if not zone and not desc:
            continue
        # 排除 noise (footer/total/banner rows)
        if zone.upper() in ("TOTAL", "SUBTOTAL", "REMARKS", "NOTE", "REMARK"):
            continue
        # 至少 zone 或 desc 一個有實質內容
        if len(zone) < 2 and len(desc) < 2:
            continue

        # ISO inference
        inline_match = ISO_RE.search(desc.upper() + " " + zone.upper())
        inferred_isos = _infer_iso_from_description(desc)

        fact = {
            "zone": zone[:120],
            "description": desc[:300],
            "row_index": i,
        }
        if inline_match:
            fact["iso_inline"] = inline_match.group(1)
        if inferred_isos:
            fact["iso_inferred"] = inferred_isos
            fact["iso"] = inferred_isos[0]  # primary
        elif inline_match:
            fact["iso"] = inline_match.group(1)

        facts.append(fact)
    return facts


def _parse_measurement_chart_sheet(sheet_rows: list[list]) -> list[dict]:
    """從 MC sheet (尺寸表) 抽 POM rows.

    支援 4 種 brand layout:
      WMT 全碼尺寸表  : POM name @ col 0, TOL @ col 4, sizes XS (0-2) / S (4-6) / M (8-10)...
      SAN POM         : POM code @ col 0, desc @ col 1, Tol+/- @ col 2, sizes XS\n0-2 / S\n4-6...
      NET 尺寸表      : POM name (中文) @ col 0, sizes S/M/L/XL/2XL/3XL @ col 2-7
      QCE KUK*        : POM Code @ col 0, desc @ col 1, 翻譯 @ col 2, Tol- @ col 4, Tol+ @ col 5, sizes 2T/3T... @ col 8+

    Strategy: 找含 size token (XS/S/M/L/XL... 或 2T/3T 或 2-20 數字) 的 header row,
    然後逐 row 抽 POM code/name/tol/sizes.
    """
    SIZE_TOKENS = {
        # Alpha
        "XS", "S", "M", "L", "XL", "XXL", "2XL", "3XL", "4XL", "5XL", "6XL", "XXS",
        # Toddler / Youth
        "2T", "3T", "4T", "5T", "6T", "YXS", "YSM", "YMD", "YLG", "YXL",
        # Numeric
        "0", "2", "4", "6", "7", "8", "10", "12", "14", "16", "18", "20", "22", "24", "26",
        "000", "00", "3", "5", "9", "11", "13", "15",
    }

    def _norm_size(s: str) -> str:
        """Extract canonical size token from cell value like 'XS (0-2)' → 'XS' or 'S\n4-6' → 'S'."""
        if not s:
            return ""
        # Try first word/token (split by space or newline)
        for sep in ("\n", " ", "(", "/"):
            if sep in s:
                s = s.split(sep)[0]
                break
        s = s.strip().upper()
        return s if s in SIZE_TOKENS else ""

    # Find header row: ≥3 size tokens + at least one POM/Points/項目 keyword
    POM_KW = ("POM", "POINTS OF MEASURE", "DESCRIPTION", "項目", "TOL")
    header_row_idx = -1
    header = None
    for i, row in enumerate(sheet_rows[:40]):
        if not row: continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        upper_blob = " ".join(c.upper() for c in cells)
        if not any(kw in upper_blob for kw in POM_KW):
            continue
        size_hits = sum(1 for c in cells if _norm_size(c))
        if size_hits >= 3:
            header_row_idx = i
            header = cells
            break

    if header_row_idx < 0:
        return []

    # Identify columns
    pom_code_col = -1
    desc_col = -1
    notes_col = -1
    tol_col = -1       # generic Tol (e.g. WMT, NET have 1 tol col)
    tol_neg_col = -1
    tol_pos_col = -1
    cn_trans_col = -1  # 翻譯 / Chinese translation
    size_cols: list[tuple[int, str]] = []

    for ci, h in enumerate(header):
        if not h:
            continue
        h_upper = h.upper().replace("\n", " ")
        # POM code/name column (first hit)
        if pom_code_col < 0:
            if h_upper in ("POM", "POM CODE", "POM #", "POINTS OF MEASURE:", "項目"):
                pom_code_col = ci
                continue
            if h_upper.startswith("POINTS OF MEASURE") or "POINT OF MEASURE" in h_upper:
                pom_code_col = ci
                continue
        # Description
        if desc_col < 0 and pom_code_col >= 0 and ci > pom_code_col:
            if ("DESCRIPTION" in h_upper or h_upper == "DESCRIPTION"):
                desc_col = ci
                continue
        # NOTES
        if notes_col < 0 and "NOTES" in h_upper:
            notes_col = ci
            continue
        # 中文翻譯
        if cn_trans_col < 0 and "翻譯" in h:
            cn_trans_col = ci
            continue
        # Tol columns
        if "TOL" in h_upper:
            # Distinguish neg vs pos
            if "-" in h and tol_neg_col < 0:
                tol_neg_col = ci
                continue
            if "+" in h and tol_pos_col < 0:
                tol_pos_col = ci
                continue
            if tol_col < 0:
                tol_col = ci
                continue
        # Size columns (token detection)
        size_name = _norm_size(h)
        if size_name:
            size_cols.append((ci, size_name))

    # Dedupe size_cols by name (QCE 等 brand 在同 row 有重複 size labels)
    # 保留 first occurrence only
    _seen_sz = set()
    _dedup = []
    for ci, name in size_cols:
        if name in _seen_sz:
            continue
        _seen_sz.add(name)
        _dedup.append((ci, name))
    size_cols = _dedup

    # WMT/NET fallback: if no POM marker but has size cols + first col looks like description
    if pom_code_col < 0 and size_cols:
        pom_code_col = 0

    if pom_code_col < 0 or not size_cols:
        return []

    # Parse data rows
    poms = []
    EMPTY_VAL = {"0", "no", "NO", "-", "", "None", "N/A", "n/a"}
    for row in sheet_rows[header_row_idx + 1:]:
        if not row: continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        if len(cells) <= pom_code_col: continue
        code = cells[pom_code_col]
        if not code or code in EMPTY_VAL: continue
        # Skip section headers (e.g. "WAIST" 行, 全空 + size col 全 0)
        # Detect: if all size cells are empty or "0", probably a section header → still keep as POM
        # but tag _no_values
        pom = {"POM_Code": code[:80]}

        # Description (English)
        if desc_col >= 0 and desc_col < len(cells):
            d = cells[desc_col]
            if d and d not in EMPTY_VAL:
                pom["POM_Name"] = d[:200]

        # 中文翻譯
        if cn_trans_col >= 0 and cn_trans_col < len(cells):
            v = cells[cn_trans_col]
            if v and v not in EMPTY_VAL:
                pom["POM_Name_zh"] = v[:120]

        # NOTES
        if notes_col >= 0 and notes_col < len(cells):
            v = cells[notes_col]
            if v and v not in EMPTY_VAL:
                pom["notes"] = v[:120]

        # Tolerance
        tol = {}
        if tol_neg_col >= 0 and tol_neg_col < len(cells):
            v = cells[tol_neg_col]
            if v and v not in EMPTY_VAL: tol["neg"] = v[:20]
        if tol_pos_col >= 0 and tol_pos_col < len(cells):
            v = cells[tol_pos_col]
            if v and v not in EMPTY_VAL: tol["pos"] = v[:20]
        if not tol and tol_col >= 0 and tol_col < len(cells):
            v = cells[tol_col]
            if v and v not in EMPTY_VAL: tol["range"] = v[:20]
        if tol:
            pom["tolerance"] = tol

        # Sizes
        sizes_dict = {}
        for col_i, size_name in size_cols:
            if col_i < len(cells):
                v = cells[col_i]
                if v and v not in EMPTY_VAL:
                    sizes_dict[size_name] = v[:30]
        if sizes_dict:
            pom["sizes"] = sizes_dict

        # Skip if neither sizes nor any other useful field
        if not (sizes_dict or tol or pom.get("POM_Name") or pom.get("POM_Name_zh")):
            continue

        poms.append(pom)

    return poms


def _worker_extract(folder_path_str: str) -> dict:
    folder = Path(folder_path_str)
    meta = _parse_folder_name(folder.name)
    facets = {
        "eidh": meta["eidh"],
        "client_code": meta["client_code"],
        "client_raw": meta["client_raw"],
        "design_id": meta["design_id"],
        "xlsx_files": [],
        "metadata": {},                # 2026-05-12 加: WMT/SAN/QCE/NET XLSX 主源 metadata
        "construction_iso_map": [],
        "measurement_charts": [],
        "_status": "ok",
    }

    xlsx_files = sorted(folder.glob("*.xlsx"))
    if not xlsx_files:
        facets["_status"] = "no_xlsx"
        return facets

    try:
        import openpyxl
    except ImportError:
        facets["_status"] = "no_openpyxl"
        return facets

    for xlsx in xlsx_files:
        facets["xlsx_files"].append(xlsx.name)
        try:
            wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
        except Exception:
            continue

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                # 讀整 sheet rows (cap to 200 to avoid huge)
                rows = list(ws.iter_rows(max_row=200, values_only=True))
            except Exception:
                continue
            # 2026-05-12 加: 試抽 metadata (cover sheet 含 STYLE/CUSTOMER/SEASON 等)
            # 只跑 WMT/SAN/QCE/NET 4 brand (XLSX 主源, PDF 無 meta)
            if facets["client_code"] in ("WMT", "SAN", "QCE", "NET") and not facets["metadata"]:
                sheet_meta = _parse_metadata_sheet(sheet_name, rows, facets["client_code"])
                if sheet_meta:
                    # Filter out 空值
                    sheet_meta = {k: v for k, v in sheet_meta.items() if v and str(v).strip() not in ("0", "")}
                    if sheet_meta:
                        sheet_meta["_source_xlsx"] = xlsx.name
                        sheet_meta["_source_sheet"] = sheet_name
                        facets["metadata"] = sheet_meta

            sheet_type = _classify_sheet(sheet_name, rows)
            if sheet_type == "translation":
                callouts = _parse_translation_sheet(rows)
                for c in callouts:
                    c["_source_xlsx"] = xlsx.name
                    c["_source_sheet"] = sheet_name
                facets["construction_iso_map"].extend(callouts)
            elif sheet_type == "measurement":
                # 2026-05-12: 結構化 POM rows (generic 4-brand parser)
                poms = _parse_measurement_chart_sheet(rows)
                if poms:
                    facets["measurement_charts"].append({
                        "_source_xlsx": xlsx.name,
                        "_source_sheet": sheet_name,
                        "n_poms": len(poms),
                        "poms": poms,
                    })
                else:
                    # 偵測到 MC sheet 但找不到 POM table (可能 sheet 結構非預期)
                    facets["measurement_charts"].append({
                        "_source_xlsx": xlsx.name,
                        "_source_sheet": sheet_name,
                        "_unparsed": True,
                    })

    return facets


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--limit", type=int, default=0)
    p.add_argument("--workers", type=int, default=4)
    p.add_argument("--client", help="只跑特定 brand code")
    p.add_argument("--reset", action="store_true", help="清舊 jsonl 重抽")
    args = p.parse_args()

    manifest_lookup = _load_manifest_lookup()
    print(f"[manifest] loaded {len(manifest_lookup):,} EIDH lookup entries")
    global _MANIFEST_LOOKUP
    _MANIFEST_LOOKUP = manifest_lookup

    folders = sorted(d for d in TP_DIR.iterdir() if d.is_dir())
    if args.client:
        folders = [f for f in folders if _parse_folder_name(f.name, manifest_lookup)["client_code"] == args.client]
    if args.limit:
        folders = folders[:args.limit]
    print(f"[scan] {TP_DIR}: {len(folders)} EIDH folders")

    if args.reset and OUT_JSONL.exists():
        OUT_JSONL.unlink()
        print(f"[reset] removed {OUT_JSONL}")

    t0 = time.time()
    stats = Counter()
    by_client = Counter()
    by_client_with_callouts = Counter()
    total_callouts = 0

    with open(OUT_JSONL, "w", encoding="utf-8") as fout, \
         ProcessPoolExecutor(max_workers=args.workers,
                             initializer=_init_worker,
                             initargs=(manifest_lookup,)) as ex:
        futures = {ex.submit(_worker_extract, str(d)): d.name for d in folders}
        for i, fut in enumerate(as_completed(futures)):
            try:
                r = fut.result()
            except Exception as e:
                print(f"  [!] {futures[fut]}: {e}", file=sys.stderr)
                stats["worker_err"] += 1
                continue
            stats[r.get("_status", "?")] += 1
            cl = r.get("client_code", "UNKNOWN")
            by_client[cl] += 1
            if r.get("construction_iso_map"):
                by_client_with_callouts[cl] += 1
                total_callouts += len(r["construction_iso_map"])
            fout.write(json.dumps(r, ensure_ascii=False) + "\n")
            if (i + 1) % 50 == 0:
                rate = (i + 1) / max(time.time() - t0, 0.1)
                eta = (len(folders) - i - 1) / rate / 60
                print(f"  [{i+1}/{len(folders)}] rate={rate:.1f}/s ETA={eta:.0f}min", flush=True)

    elapsed_min = (time.time() - t0) / 60
    print(f"\n[done] {sum(stats.values())} folders in {elapsed_min:.1f} min")
    print(f"  total construction_iso_map: {total_callouts:,}")
    print(f"\nstatus:")
    for s, n in stats.most_common():
        print(f"  {s:<15} {n:>6}")
    print(f"\nby client (total / w/construction_iso_map):")
    for cl in sorted(by_client.keys(), key=lambda x: -by_client[x]):
        print(f"  {cl:<8} {by_client[cl]:>5} {by_client_with_callouts[cl]:>5}")
    print(f"\noutput: {OUT_JSONL}")


if __name__ == "__main__":
    main()
