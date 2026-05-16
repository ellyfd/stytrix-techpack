#!/usr/bin/env python3
"""build_customer_terminology.py — 蒸餾 customer glossary → runtime JSON (2026-05-16)

Source: Source-Data/做工翻譯/ISO_客人縫法對照表_Glossary.xlsx
  Sheet1 "ISO x 客人縫法對照" — 13 客人 / 31,375 筆 / 11 ISO codes
  Sheet2 "客人 Style Profile" — 13 客人 × 7 attributes

Output:
  data/runtime/customer_terminology_master.json — Layer 1: ISO + 客人 phrasings
  data/runtime/customer_style_profile.json      — Layer 2: 各客人 style profile

跑法（在 PowerShell，repo cwd）:
  python scripts\\core\\build_customer_terminology.py

如果 raw xlsx 不在標準位置可加 --source:
  python scripts\\core\\build_customer_terminology.py --source "C:\\path\\to\\Glossary.xlsx"

Future: l5_anchors 待補 — 等 iso_lookup_5dim machine→L5 reverse map 接通後 enrich entry
"""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl not installed. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_SOURCE = Path(r"C:\Users\ellycheng.DOMAIN\Desktop\StyTrix TechPack\Construction\Source-Data\做工翻譯\ISO_客人縫法對照表_Glossary.xlsx")
ROOT = Path(__file__).resolve().parent.parent.parent
OUT_DIR = ROOT / "data" / "runtime"


def build_master(xlsx_path: Path):
    """Build Layer 1: ISO-centric mapping with 13 客人 phrasings."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["ISO x 客人縫法對照"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 1 = client headers ("DKS\n7,191筆" / "KOH\n5,649筆" / ...)
    header = rows[1]
    clients = []
    for cell in header[2:]:
        if cell:
            code = cell.split("\n")[0].strip()
            clients.append(code)

    # Pairs from row 2 onwards: phrasings row + counts row
    entries = []
    for i in range(2, len(rows), 2):
        if i + 1 >= len(rows):
            break
        phr_row = rows[i]
        cnt_row = rows[i + 1]
        iso = phr_row[0]
        zh = phr_row[1]
        if not iso:
            continue
        iso = str(iso).strip()
        zh_clean = (zh.replace("\n", " / ") if zh else "").strip()

        customer_variants = {}
        for ci, client in enumerate(clients):
            phr_cell = phr_row[2 + ci]
            cnt_cell = cnt_row[2 + ci]
            if phr_cell:
                phrasings = [p.strip() for p in str(phr_cell).split("\n") if p.strip()]
                customer_variants[client] = {
                    "phrasings": phrasings,
                    "count": int(cnt_cell) if cnt_cell else 0,
                    "primary": phrasings[0] if phrasings else None,
                }

        entries.append({
            "canonical": {
                "iso": iso,
                "machine_zh": zh_clean,
                "l5_anchors": [],
            },
            "customer_variants": customer_variants,
        })

    wb.close()

    return {
        "_version": "v1",
        "_generated_at_source": str(xlsx_path.name),
        "_description": "客人縫法 master mapping — ISO 為中心,各客人寫法掛旁邊。",
        "_consumer": [
            "ingest 端: 客人 phrasing → ISO code (reverse map)",
            "output 端: 五階 → ISO → 客人慣用語 (forward map)",
        ],
        "_canonical_clients": clients,
        "_todo": "l5_anchors 待補 — 從 iso_lookup_5dim machine→L5 reverse + iso_dictionary 接通",
        "entries": entries,
    }


def build_profile(xlsx_path: Path):
    """Build Layer 2: per-客人 style profile (給 output formatter)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["客人 Style Profile"]
    rows = list(ws.iter_rows(values_only=True))

    clients_profile = {}
    for r in rows[1:]:
        if not r[0]:
            continue
        code = str(r[0]).strip()
        rate_str = str(r[2] or "").replace("%", "").strip()
        iso_rate = float(rate_str) / 100.0 if rate_str and rate_str.replace(".", "").isdigit() else 0.0
        clients_profile[code] = {
            "count": int(r[1] or 0),
            "iso_label_rate": iso_rate,
            "format_type": str(r[3] or "").strip(),
            "primary_406_variant": str(r[4] or "").strip(),
            "default_stitch": str(r[5] or "").strip(),
            "identifying_features": str(r[6] or "").strip(),
        }
    wb.close()

    return {
        "_version": "v1",
        "_generated_at_source": str(xlsx_path.name),
        "_description": "各客人 縫法書寫風格 profile — 給做工建議 output formatter 用。",
        "_consumer": "前端做工建議 formatter (五階 L5 → ISO → 客人 phrasing 時套客人 format)",
        "_field_descriptions": {
            "iso_label_rate": "客人 PDF 中標準 ISO 編號的比例 (0.21 = 21%)",
            "format_type": "客人寫縫法的格式類型 (直標ISO / 縮寫+ISO / 功能描述 / 不使用)",
            "primary_406_variant": "客人寫 ISO 406 (三本雙針) 的最常用變體寫法",
            "default_stitch": "客人 預設縫法 (PDF 開頭聲明 或 不聲明)",
            "identifying_features": "客人寫法的識別特徵 (大寫率 / 特殊術語 / 等)",
        },
        "clients": clients_profile,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", default=str(DEFAULT_SOURCE), help="Glossary xlsx path")
    args = ap.parse_args()

    xlsx_path = Path(args.source)
    if not xlsx_path.exists():
        print(f"[ERROR] source not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"[build_customer_terminology] reading {xlsx_path}")

    master = build_master(xlsx_path)
    profile = build_profile(xlsx_path)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    master_path = OUT_DIR / "customer_terminology_master.json"
    master_path.write_text(json.dumps(master, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] {master_path} written: {len(master['entries'])} ISO codes × {len(master['_canonical_clients'])} clients")

    profile_path = OUT_DIR / "customer_style_profile.json"
    profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] {profile_path} written: {len(profile['clients'])} clients")


if __name__ == "__main__":
    main()
