"""report_l5_audit_typed.py — L4 mismatch 拆 Type A/B/C 給 IE 部門看

讀 m7_pullon_designs.jsonl + Bible(stytrix-techpack/l2_l3_ie/<L1>.json 38 檔),
把每筆 mismatch step 標 Type:
  A — IE 端 broad category(命名顆粒度差異,不是 Bible 缺結構)
  B — Placeholder(per CLAUDE.md 該 drop)
  C — `*_其它` 結尾(IE 端「未細分」placeholder)
  other — 真實新結構(IE 部門要看是否補進 Bible)

輸出:
  console — overall + per Type breakdown
  outputs/platform/l5_audit_typed.txt — 完整 console 報告
  outputs/platform/l5_audit_typed.xlsx — IE 部門可篩選版,sheet:
    Sheet 1 (summary):  per Type 件數 / unique L4 數 / per L1 / per Brand
    Sheet 2 (Type B placeholders):  IE 要 finalize 的 new_*  list,含 EIDH / count
    Sheet 3 (Type other 真實新結構):  Bible 缺什麼,IE 部門要補

跑:python scripts/report_l5_audit_typed.py [--bible-root <path>]
"""
from __future__ import annotations
import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts" / "lib"))
from bible_classify import load_bible, step_alignment  # noqa: E402

DESIGNS = ROOT / "outputs" / "platform" / "m7_pullon_designs.jsonl"
DEFAULT_BIBLE_ROOT = Path(r"C:\temp\stytrix-techpack\l2_l3_ie")
OUT_TXT = ROOT / "outputs" / "platform" / "l5_audit_typed.txt"
OUT_XLSX = ROOT / "outputs" / "platform" / "l5_audit_typed.xlsx"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--bible-root", default=str(DEFAULT_BIBLE_ROOT))
    args = ap.parse_args()

    bible = load_bible(Path(args.bible_root))
    print(f"[bible] loaded {bible['n_files']} L1 files / {len(bible['full_tuples']):,} canonical tuples")

    if not DESIGNS.exists():
        print(f"[FAIL] {DESIGNS} not found")
        return 1

    # === Walk designs,分類每筆 step ===
    print(f"[walk] {DESIGNS.relative_to(ROOT)}")
    n_steps = 0
    type_counter = Counter()
    per_l1_type = defaultdict(Counter)
    per_brand_type = defaultdict(Counter)
    # Type-specific evidence collectors
    type_a_l4 = Counter()        # broad category 對應 L4 名 → count
    type_b_placeholder = []       # list of (eidh, brand, l4)
    type_c_other_l4 = Counter()
    type_other_l4 = Counter()    # 真新結構 → count
    type_other_evidence = defaultdict(set)  # (l1, wk, l2, l3, l4) → set of EIDH

    n_designs = 0
    with open(DESIGNS, encoding="utf-8") as f:
        for line in f:
            try:
                d = json.loads(line)
            except Exception:
                continue
            n_designs += 1
            brand = d.get("client", {}).get("code", "?")
            fabric = (d.get("fabric") or {}).get("value", "").lower()
            eidh = str(d.get("eidh", ""))
            steps = d.get("five_level_steps") or []

            for s in steps:
                n_steps += 1
                align = step_alignment(s, fabric, bible)
                t = align["type"]
                type_counter[t] += 1
                l1 = s.get("l1", "?")
                per_l1_type[l1][t] += 1
                per_brand_type[brand][t] += 1

                if t == "match":
                    continue

                l4 = s.get("l4", "")
                wk = fabric
                key = (l1, wk, s.get("l2", ""), s.get("l3", ""), l4)

                if t == "A":
                    type_a_l4[l4] += 1
                elif t == "B":
                    type_b_placeholder.append({
                        "eidh": eidh, "brand": brand, "l1": l1, "wk": wk,
                        "l2": s.get("l2", ""), "l3": s.get("l3", ""),
                        "l4": l4, "l5": s.get("l5", ""),
                    })
                elif t == "C":
                    type_c_other_l4[l4] += 1
                elif t == "other":
                    type_other_l4[l4] += 1
                    type_other_evidence[key].add(eidh)

    n_match = type_counter["match"]
    n_mismatch = n_steps - n_match

    # === Console + txt 輸出 ===
    lines = []
    def emit(s=""): print(s); lines.append(s)

    emit("=" * 110)
    emit(f"L5 Audit (Typed) Report")
    emit(f"  source:  {DESIGNS.relative_to(ROOT)}")
    emit(f"  bible:   {args.bible_root}")
    emit("=" * 110)
    emit(f"\n[Overall]")
    emit(f"  n_designs:                 {n_designs:>9,}")
    emit(f"  n_steps:                   {n_steps:>9,}")
    emit(f"  match:                     {n_match:>9,}  ({100*n_match/n_steps:.1f}%)")
    emit(f"  mismatch (total):          {n_mismatch:>9,}  ({100*n_mismatch/n_steps:.1f}%)")
    emit(f"")
    emit(f"  Type A (broad category):   {type_counter['A']:>9,}  ({100*type_counter['A']/n_steps:>5.1f}%)  ← IE 端命名,非 Bible bug")
    emit(f"  Type B (placeholder NEW_): {type_counter['B']:>9,}  ({100*type_counter['B']/n_steps:>5.1f}%)  ← IE 要 finalize,該 drop 不該進 Bible")
    emit(f"  Type C (*_其它 結尾):      {type_counter['C']:>9,}  ({100*type_counter['C']/n_steps:>5.1f}%)  ← IE「未細分」placeholder,可保留")
    emit(f"  Type other (真實新結構):   {type_counter['other']:>9,}  ({100*type_counter['other']/n_steps:>5.1f}%)  ← Bible 缺,IE 要看")

    emit(f"\n[per L1 — Type breakdown]")
    emit(f"  {'L1':5}  {'total':>7}  {'match':>7}  {'A':>6}  {'B':>5}  {'C':>5}  {'other':>5}")
    sorted_l1 = sorted(per_l1_type.keys(), key=lambda l: -sum(per_l1_type[l].values()))
    for l1 in sorted_l1[:30]:
        c = per_l1_type[l1]
        tot = sum(c.values())
        emit(f"  {l1:5}  {tot:>7,}  {c['match']:>7,}  {c['A']:>6,}  {c['B']:>5,}  {c['C']:>5,}  {c['other']:>5,}")

    emit(f"\n[per Brand — Type breakdown]")
    emit(f"  {'brand':10}  {'total':>7}  {'match':>7}  {'A':>6}  {'B':>5}  {'C':>5}  {'other':>5}")
    for brand in sorted(per_brand_type.keys(), key=lambda b: -sum(per_brand_type[b].values())):
        c = per_brand_type[brand]
        tot = sum(c.values())
        emit(f"  {brand:10}  {tot:>7,}  {c['match']:>7,}  {c['A']:>6,}  {c['B']:>5,}  {c['C']:>5,}  {c['other']:>5,}")

    emit(f"\n[Type B — Placeholder 樣本 (前 20 件,IE 要 finalize)]")
    # group by L4 placeholder name
    b_by_l4 = Counter(p["l4"] for p in type_b_placeholder)
    for l4, n in b_by_l4.most_common(20):
        # 找一個 sample EIDH
        sample = next((p for p in type_b_placeholder if p["l4"] == l4), None)
        if sample:
            emit(f"  {n:>5}× {sample['l1']}/{sample['wk']}/{sample['l2']}/{sample['l3']}/[{l4}]"
                 f" (sample EIDH={sample['eidh']})")

    emit(f"\n[Type other — 真新結構 Top 20 (IE 要看是否補進 Bible)]")
    sorted_other = sorted(type_other_evidence.items(),
                          key=lambda kv: -type_other_l4[kv[0][4]])[:20]
    for key, eidh_set in sorted_other:
        l1, wk, l2, l3, l4 = key
        n = type_other_l4[l4]
        emit(f"  {n:>5}× {l1}/{wk}/{l2}/{l3}/[{l4[:50]}]  涉及 {len(eidh_set)} EIDH")

    OUT_TXT.parent.mkdir(parents=True, exist_ok=True)
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n[output] {OUT_TXT}")

    # === xlsx ===
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[skip xlsx] openpyxl not installed — pip install openpyxl")
        return 0

    wb = Workbook()
    # Sheet 1 — summary
    ws1 = wb.active
    ws1.title = "summary"
    ws1.append(["指標", "件數", "占比%", "說明"])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for label, n, note in [
        ("Total steps",      n_steps,            ""),
        ("Match (對得上 Bible)", n_match,        "L1/L2/L3/L4/L5 全 OK"),
        ("Type A broad cat", type_counter["A"], "手工類/車縫類/打結 等大分類詞 — IE 端合理命名,非 Bible bug"),
        ("Type B placeholder", type_counter["B"], "new_method_describe_* / (NEW)* — IE 要 finalize,Phase 2 該 drop"),
        ("Type C *_其它",    type_counter["C"], "L4 結尾 _其它 — IE「未細分」placeholder,可保留"),
        ("Type other",       type_counter["other"], "真實新結構 — Bible 缺,IE 部門要看是否補"),
    ]:
        pct = 100 * n / n_steps if n_steps else 0
        ws1.append([label, n, round(pct, 2), note])

    # Sheet 2 — per L1 × Type
    ws_l1 = wb.create_sheet("per_L1")
    ws_l1.append(["L1", "L1_zh", "total", "match", "Type_A", "Type_B", "Type_C", "Type_other", "match_pct"])
    for cell in ws_l1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    l1_zh = bible["l1_zh"]
    for l1 in sorted_l1:
        c = per_l1_type[l1]
        tot = sum(c.values())
        match_pct = round(100 * c["match"] / tot, 1) if tot else 0
        ws_l1.append([l1, l1_zh.get(l1, ""), tot, c["match"],
                      c["A"], c["B"], c["C"], c["other"], match_pct])

    # Sheet 3 — Type B placeholders (IE 要 finalize)
    ws_b = wb.create_sheet("Type_B_placeholder_finalize")
    ws_b.append(["L4_placeholder", "count", "L1", "wk", "L2", "L3",
                 "sample_EIDH", "sample_L5", "sample_brand"])
    for cell in ws_b[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFE4B5")  # 橘色提示
    for l4, n in b_by_l4.most_common():
        sample = next((p for p in type_b_placeholder if p["l4"] == l4), None)
        if sample:
            ws_b.append([l4, n, sample["l1"], sample["wk"], sample["l2"],
                         sample["l3"], sample["eidh"], sample["l5"], sample["brand"]])

    # Sheet 4 — Type other (Bible 真缺結構)
    ws_o = wb.create_sheet("Type_other_bible_gaps")
    ws_o.append(["L1", "wk", "L2", "L3", "L4", "count_steps", "n_EIDH",
                 "L1_zh", "comment"])
    for cell in ws_o[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFC7CE")  # 紅色 = 真要補
    for key, eidh_set in sorted(type_other_evidence.items(),
                                key=lambda kv: -type_other_l4[kv[0][4]]):
        l1, wk, l2, l3, l4 = key
        n = type_other_l4[l4]
        ws_o.append([l1, wk, l2, l3, l4, n, len(eidh_set),
                     l1_zh.get(l1, ""), ""])

    # Sheet 5 — Type A broad cat (參考)
    ws_a = wb.create_sheet("Type_A_broad_category")
    ws_a.append(["L4_broad_token", "count_steps", "說明"])
    for cell in ws_a[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="C6EFCE")  # 綠色 = OK
    for l4, n in type_a_l4.most_common(50):
        ws_a.append([l4, n, "IE 端 broad category — 不需 Bible 改"])

    # 自動調整欄寬
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for col in sheet.columns:
            try:
                width = max(len(str(c.value or "")) for c in col)
                sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 60)
            except Exception:
                pass

    wb.save(OUT_XLSX)
    print(f"[output] {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main() or 0)
