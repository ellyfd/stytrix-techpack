"""pdf_facets_to_mc_pom.py — Phase 2 adapter: pdf_facets/xlsx_facets → mc_pom_*.jsonl

Pipeline B (pom_rules) 的 rebuild_profiles.py 讀 `$BASE/_parsed/mc_pom_*.jsonl`,
schema 跟 M7 Pipeline 的 pdf_facets.jsonl 不同 (但 poms[] 內層欄位完全一樣).

這個 adapter 把 pdf_facets + xlsx_facets 內所有「有 POM」的 entry 轉成
mc_pom 格式, 讓 Pipeline B 能吃到救出的 611K POMs.

mc_pom schema (rebuild_profiles.py 期待):
  {
    "design_number": str,
    "brand_division": str,
    "department": str,
    "item_type": str,
    "manifest_item": str,   # 聚陽 M7列管 Item — garment_type
    "mk_gender": str,       # 聚陽 M7列管 PRODUCT_CATEGORY → gender code
    "mk_fabric": str,       # 聚陽 M7列管 W/K → Knit/Woven
    "category": str,
    "design_type": str,
    "description": str,
    "mcs": [ {body_type, sizes:[...], poms:[{POM_Code,POM_Name,sizes,tolerance}]} ]
  }

跑法:
  python scripts/pdf_facets_to_mc_pom.py [--out PATH] [--brands ONY GAP ...] [--include-xlsx]

預設 --out = outputs/platform/mc_pom_v11.jsonl
跑完手動 copy 到 $POM_PIPELINE_BASE/_parsed/ 再跑 Pipeline B Step 2-6.
"""
import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
EXTRACT_DIR = ROOT / "outputs" / "extract"
PDF_FACETS = EXTRACT_DIR / "pdf_facets.jsonl"
XLSX_FACETS = EXTRACT_DIR / "xlsx_facets.jsonl"

# 2026-05-14 Elly: 三個維度都對映到聚陽 M7列管 (M7列管_*.xlsx 的「總表」sheet),
# 不再讓各 brand 用自己的 metadata 各自亂猜 GT/gender/fabric:
#   Item             → manifest_item  (garment_type, 聚陽款式原值)
#   PRODUCT_CATEGORY → mk_gender       (Women/Men/Girl/Boy/Baby → 性別碼)
#   W/K              → mk_fabric       (Knit/Woven)
# 全部用 EIDH 反查, 100% 對得到。M7列管 在 Source-Data/ 下 (adapter 的上一層)。
# 找不到 M7列管 時 fallback 到 _fetch_manifest.csv (只有 Item, 無 gender/fabric)。
_MK = {}  # eidh(str) -> {"item": str, "gender": str, "fabric": str}

_PC2GENDER = {
    "Women": "WOMENS", "Men": "MENS", "Girl": "GIRLS",
    "Boy": "BOYS", "Baby": "BABY/TODDLER",
}


def _norm_gender(pc):
    """M7列管 PRODUCT_CATEGORY (e.g. 'Women 女士') → gender code."""
    pc = (pc or "").strip()
    for en, code in _PC2GENDER.items():
        if pc.startswith(en):
            return code
    return ""


def _load_manifest_csv_fallback():
    """舊路徑: _fetch_manifest.csv 只有 Item, gender/fabric 留空."""
    csv_path = ROOT / "m7_organized_v2" / "_fetch_manifest.csv"
    if not csv_path.exists():
        print(f"[!] manifest 也不存在: {csv_path} — manifest_item/mk_gender/mk_fabric 全空")
        return
    import csv as _csv
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in _csv.DictReader(f):
            eidh = str(row.get("Eidh", "")).strip()
            item = (row.get("Item", "") or "").strip()
            if eidh and item:
                _MK[eidh] = {"item": item, "gender": "", "fabric": ""}
    print(f"[manifest CSV fallback] {len(_MK):,} eidh→Item (無 gender/fabric)")


def load_mk_control():
    """讀 M7列管_*.xlsx 的「總表」sheet, 建 eidh → {item, gender, fabric}."""
    candidates = sorted(ROOT.parent.glob("M7列管_*.xlsx"))
    if not candidates:
        print(f"[!] 找不到 M7列管_*.xlsx (在 {ROOT.parent}), fallback _fetch_manifest.csv")
        _load_manifest_csv_fallback()
        return
    ctrl = candidates[-1]  # 取檔名最新的 (日期戳排序)
    try:
        import openpyxl
    except ImportError:
        print("[!] openpyxl 未安裝, 無法讀 M7列管, fallback _fetch_manifest.csv")
        _load_manifest_csv_fallback()
        return
    wb = openpyxl.load_workbook(ctrl, read_only=True, data_only=True)
    if "總表" not in wb.sheetnames:
        print(f"[!] M7列管 沒有「總表」sheet ({wb.sheetnames}), fallback CSV")
        wb.close()
        _load_manifest_csv_fallback()
        return
    ws = wb["總表"]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}
    need = ("Eidh", "Item", "PRODUCT_CATEGORY", "W/K")
    missing = [c for c in need if c not in idx]
    if missing:
        print(f"[!] M7列管 總表 缺欄位 {missing}, fallback CSV")
        wb.close()
        _load_manifest_csv_fallback()
        return
    for r in rows:
        eidh = r[idx["Eidh"]]
        if eidh is None:
            continue
        eidh = str(eidh).strip()
        _MK[eidh] = {
            "item": (str(r[idx["Item"]]).strip() if r[idx["Item"]] is not None else ""),
            "gender": _norm_gender(r[idx["PRODUCT_CATEGORY"]]),
            "fabric": (str(r[idx["W/K"]]).strip() if r[idx["W/K"]] is not None else ""),
        }
    wb.close()
    print(f"[M7列管] {len(_MK):,} eidh 載入 ({ctrl.name}) "
          f"— Item + PRODUCT_CATEGORY(gender) + W/K(fabric)")


# 2026-05-13 Elly: 小量混入的 dev / 雜項 brand 不進 pom_rules
# (V2 DEV / V5 DEV = 內部開發線, JF/ROSS/ZAR = 雜項小客戶, 各 1-3 designs)
# pom_rules 是通用尺寸規則庫, 小量雜訊 brand 會污染 bucket 統計
DROP_BRANDS = {"V2 DEV", "V5 DEV", "V7 DEV", "S1 DEV", "JF", "ROSS", "ZAR",
               "HLA", "DST", "ASICS", "GILDAN", "BUSINE", "CLUB M",
               "CARTER", "CATAPU", "AEOM", "QUINCE"}  # QUINCE → QCE 才對, 排重複碼


def _inject_gender(brand_division, metadata):
    """fallback: M7列管 對不到 eidh 時, 把 metadata 的 gender 塞進 brand_division 末尾,
    讓 rebuild_profiles.py 的 extract_gender() 還抓得到 (主路徑是 mk_gender)."""
    bd = brand_division or ""
    bd_upper = bd.upper()
    for g in ("GIRLS", "BOYS", "WOMENS", "MENS", "MATERNITY", "BABY", "TODDLER"):
        if g in bd_upper:
            return bd
    g = (metadata.get("gender") or metadata.get("gender_inferred")
         or metadata.get("gender_raw") or "").upper().strip()
    g_map = {
        "WOMEN": "WOMENS", "WOMEN'S": "WOMENS", "FEMALE": "WOMENS",
        "MEN": "MENS", "MEN'S": "MENS", "MALE": "MENS",
        "GIRL": "GIRLS", "BOY": "BOYS",
        "KIDS": "", "YOUTH": "", "UNISEX": "",
    }
    g = g_map.get(g, g)
    if g in ("GIRLS", "BOYS", "WOMENS", "MENS", "MATERNITY"):
        return f"{bd} {g}".strip()
    class_raw = (metadata.get("class_raw") or "").upper()
    for gtok in ("GIRLS", "BOYS", "WOMENS", "MENS"):
        if gtok in class_raw:
            return f"{bd} {gtok}".strip()
    return bd


def _convert_entry(e):
    """pdf_facets / xlsx_facets entry → mc_pom record. 無 POM 回 None."""
    mcs_in = e.get("measurement_charts") or []
    mcs_out = []
    for mc in mcs_in:
        poms = mc.get("poms") or []
        if not poms:
            continue
        mcs_out.append({
            "body_type": mc.get("body_type", "") or mc.get("base_size", ""),
            "sizes": mc.get("sizes") or [],
            "poms": poms,   # POM_Code/POM_Name/sizes/tolerance 欄位完全一樣, 直接帶
        })
    if not mcs_out:
        return None

    meta = e.get("metadata") or {}
    brand_division = meta.get("brand_division") or e.get("client_raw") or e.get("client_code") or ""
    brand_division = _inject_gender(brand_division, meta)

    eidh = str(e.get("eidh", "")).strip()
    mk = _MK.get(eidh, {})
    return {
        "design_number": (meta.get("design_number") or e.get("design_id") or "").strip(),
        "brand_division": brand_division,
        "department": meta.get("department") or meta.get("department_raw") or "",
        "item_type": meta.get("item_type") or "",
        # 2026-05-14: 三個維度都從聚陽 M7列管 反查 (EIDH key):
        #   manifest_item = Item (garment_type 原值)
        #   mk_gender     = PRODUCT_CATEGORY → gender code
        #   mk_fabric     = W/K (Knit/Woven)
        # 下游 rebuild_profiles.py / reclassify_and_rebuild.py 優先吃這三欄。
        "manifest_item": mk.get("item", ""),
        "mk_gender": mk.get("gender", ""),
        "mk_fabric": mk.get("fabric", ""),
        "category": (meta.get("bom_category") or meta.get("sub_category")
                     or meta.get("class_raw") or meta.get("category") or ""),
        "design_type": meta.get("design_type") or "",
        "description": meta.get("description") or "",
        "mcs": mcs_out,
        # audit trail
        "_eidh": eidh,
        "_client_code": e.get("client_code", ""),
        "_source": "pdf_facets" if "_source_pdf" in str(mcs_in) else "facets",
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(ROOT / "outputs" / "platform" / "mc_pom_v11.jsonl"))
    ap.add_argument("--brands", nargs="*", default=None,
                    help="只轉特定 brand code (預設全部有 POM 的)")
    ap.add_argument("--include-xlsx", action="store_true",
                    help="也轉 xlsx_facets (SAN/WMT/QCE/NET POM)")
    args = ap.parse_args()

    load_mk_control()

    target_brands = set(b.upper() for b in args.brands) if args.brands else None
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    sources = [("pdf_facets", PDF_FACETS)]
    if args.include_xlsx:
        sources.append(("xlsx_facets", XLSX_FACETS))

    n_in = 0
    n_with_pom = 0
    n_written = 0
    n_dropped = 0
    n_item_hit = 0
    n_gender_hit = 0
    n_fabric_hit = 0
    by_brand = {}
    pom_total = 0

    with open(out_path, "w", encoding="utf-8") as fout:
        for src_name, src_path in sources:
            if not src_path.exists():
                print(f"[!] {src_path} 不存在, skip")
                continue
            print(f"[read] {src_path.name}")
            with open(src_path, encoding="utf-8") as f:
                for line in f:
                    try:
                        e = json.loads(line)
                    except Exception:
                        continue
                    n_in += 1
                    cl = e.get("client_code", "?")
                    if target_brands and cl not in target_brands:
                        continue
                    if cl in DROP_BRANDS:
                        n_dropped += 1
                        continue
                    rec = _convert_entry(e)
                    if rec is None:
                        continue
                    n_with_pom += 1
                    rec["_source"] = src_name
                    if rec.get("manifest_item"):
                        n_item_hit += 1
                    if rec.get("mk_gender"):
                        n_gender_hit += 1
                    if rec.get("mk_fabric"):
                        n_fabric_hit += 1
                    fout.write(json.dumps(rec, ensure_ascii=False) + "\n")
                    n_written += 1
                    b = by_brand.setdefault(cl, {"designs": 0, "poms": 0})
                    b["designs"] += 1
                    b["poms"] += sum(len(mc["poms"]) for mc in rec["mcs"])
                    pom_total += sum(len(mc["poms"]) for mc in rec["mcs"])

    def _pct(n):
        return (100.0 * n / n_written) if n_written else 0.0
    print()
    print(f"=== Adapter Summary ===")
    print(f"  Entries scanned:       {n_in:,}")
    print(f"  Dropped (DROP_BRANDS):  {n_dropped:,}")
    print(f"  Entries with POM:      {n_with_pom:,}")
    print(f"  Written to mc_pom:     {n_written:,}")
    print(f"  manifest_item 對到:    {n_item_hit:,} ({_pct(n_item_hit):.1f}%)  ← garment_type")
    print(f"  mk_gender 對到:        {n_gender_hit:,} ({_pct(n_gender_hit):.1f}%)  ← gender")
    print(f"  mk_fabric 對到:        {n_fabric_hit:,} ({_pct(n_fabric_hit):.1f}%)  ← fabric")
    print(f"  Total POMs:            {pom_total:,}")
    print()
    print(f"  Per-brand:")
    for cl in sorted(by_brand, key=lambda c: -by_brand[c]["poms"]):
        b = by_brand[cl]
        print(f"    {cl:<8} {b['designs']:>5} designs / {b['poms']:>7,} POMs")
    print()
    print(f"  [write] {out_path}")
    print()
    print(f"=== Next step (Pipeline B) ===")
    print(f"  1. copy {out_path.name} → $POM_PIPELINE_BASE/_parsed/")
    print(f"  2. cd stytrix-techpack")
    print(f"  3. python scripts/core/rebuild_profiles.py        # Step 2 → measurement_profiles_union.json")
    print(f"  4. python scripts/core/reclassify_and_rebuild.py  # Step 3 → pom_rules/*.json")
    print(f"  5. python scripts/core/rebuild_all_analysis_v2.py # Step 4")
    print(f"  6. python scripts/core/rebuild_grading_3d.py      # Step 5")
    print(f"  7. python scripts/core/fix_sort_order.py          # Step 6")
    print(f"  8. git add pom_rules/ + commit + push → Vercel auto-deploy")
    return 0


if __name__ == "__main__":
    sys.exit(main())
