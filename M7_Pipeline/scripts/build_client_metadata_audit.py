"""build_client_metadata_audit.py — 統一輸出單一 xlsx 三 sheet

讀:
  1. m7_organized_v2/outputs/platform/pdf_cover_universe.summary.json (我們抽到的 PDF 內容)
  2. data/client_metadata_mapping.json (已有的 23 客戶 canonical mapping reference)

輸出單一 xlsx:outputs/platform/22_client_pdf_metadata_audit.xlsx

三 sheet:
  - sheet 1 client_field_universe: 22 客戶 × union of all PDF fields,coverage% 矩陣
  - sheet 2 field_value_attributes: 每 (client, field) 的所有 unique values
  - sheet 3 client_canonical_xwalk: 用 client_metadata_mapping.json 推出每 (client, field) → 聚陽 canonical

跑:python scripts/build_client_metadata_audit.py
"""
from __future__ import annotations
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SUMMARY = ROOT / "outputs" / "platform" / "pdf_cover_universe.summary.json"
MAPPING_JSON = ROOT / "data" / "client_metadata_mapping.json"
OUT_XLSX = ROOT / "outputs" / "platform" / "22_client_pdf_metadata_audit.xlsx"

MIN_COV = 30  # coverage% threshold

# user 指定 only 8 個 canonical
CANONICAL_FIELDS = [
    "客戶", "報價款號", "Program", "Subgroup", "W/K", "Item", "Season", "PRODUCT_CATEGORY",
]


def derive_canonical_from_mapping(client_pdf_name: str, field: str, top_values: list,
                                   mapping: dict) -> tuple[str, str, str]:
    """
    用 client_metadata_mapping.json 推 (client, field) → canonical
    Returns (canonical, confidence, source_in_mapping)

    Logic:
      1. 找 mapping[client] 裡每個 example field
      2. 看 field name 跟 mapping 哪個 key match (exact / fuzzy)
      3. 看 top_values 內容跟 mapping examples 是否 match
    """
    # client name 對應 (mapping uses 'ONY' / 'OLD NAVY' style)
    pdf_client_map = {
        "OLD NAVY": "ONY", "DICKS SPORTING GOODS": "DICKS",
        "GAP OUTLET": "GAP_OUTLET", "BANANA REPUBLIC": "BR",
        "UNDER ARMOUR": "UNDER_ARMOUR", "BEYOND YOGA": "BEYOND_YOGA",
        "JOE FRESH": "JOE_FRESH", "HIGH LIFE LLC": "HIGH_LIFE_LLC",
        "S1 DEVELOPING": "S1_DEVELOPING", "V5 DEVELOPING": "V5_DEVELOPING",
        "V2 DEVELOPING": "V2_DEVELOPING",
        "A & F": "A_&_F",
    }
    client_key = pdf_client_map.get(client_pdf_name, client_pdf_name)
    if client_key not in mapping.get("clients", {}):
        return "(client 不在 mapping 內)", "LOW", ""

    cm = mapping["clients"][client_key]
    fl = field.lower().strip(":：# ")
    sample_vals_lower = [str(v.get("v", "")).lower() for v in top_values[:5]]
    sample_vals_str = " ".join(sample_vals_lower)

    # === 報價款號 — 看 design_id_examples ===
    examples = [str(e).lower() for e in cm.get("design_id_examples", [])]
    formats = [str(e).lower() for e in cm.get("design_id_formats", [])]
    if examples:
        # field name keyword
        if any(kw in fl for kw in ["style #", "style#", "style no", "style code", "style number",
                                    "デザイン名", "款号", "item trading code"]):
            return "報價款號", "HIGH", f"匹配 design_id_examples: {examples[:2]}"
        # value match
        for sv in sample_vals_lower:
            if any(sv == ex or (len(sv) > 4 and ex in sv) for ex in examples[:5] if len(ex) > 4):
                return "報價款號", "HIGH", f"value match design_id_examples"

    # === 客戶 — 看 brand_division_examples + client_name_aliases ===
    bd_examples = [str(e).lower() for e in cm.get("brand_division_examples", [])]
    aliases = [str(e).lower() for e in cm.get("client_name_aliases", [])]
    full_name = str(cm.get("client_name_full", "")).lower()
    all_brand_strings = bd_examples + aliases + ([full_name] if full_name else [])
    if all_brand_strings:
        if any(kw in fl for kw in ["brand/", "brand division", "brand category",
                                    "ブランド", "品牌", "company name", "企業"]):
            return "客戶", "HIGH", f"匹配 brand_division_examples: {bd_examples[:1]}"
        if fl == "brand":
            return "客戶", "HIGH", "field=Brand"
        for sv in sample_vals_lower:
            for bs in all_brand_strings:
                if sv == bs or (bs and bs in sv):
                    return "客戶", "MEDIUM", f"value match brand"

    # === Season — 看 season_format / season_examples ===
    season_examples = [str(e).lower() for e in cm.get("season_examples", [])]
    season_format = str(cm.get("season_format", "")).lower()
    if season_examples:
        if fl in ("season", "シーズン") or "design season" in fl:
            return "Season", "HIGH", f"匹配 season_examples: {season_examples[:1]}"
        for sv in sample_vals_lower:
            if any(sv == se for se in season_examples):
                return "Season", "HIGH", "value match season_examples"
            # season pattern match (含年份 / SS/FW)
            if re.search(r"(SS|FW|AW|Fall|Spring|Summer|Holiday|Winter|20\d\d)", sv, re.I):
                if any(kw in fl for kw in ["season"]):
                    return "Season", "MEDIUM", "value 含 season pattern"

    # === Program / Subgroup — 看 subgroup_codes ===
    subgroup_codes = cm.get("subgroup_codes", {})
    if subgroup_codes:
        # value 直接在 subgroup_codes (eg 'WAC' / 'MAC' / 'WOMEN FLEECE')
        for sv in [str(v.get("v","")) for v in top_values[:5]]:
            if sv.upper() in subgroup_codes:
                return "Subgroup", "HIGH", f"value '{sv}' 在 subgroup_codes"

    # === Item — 看 item / category_examples / department_examples ===
    item_examples = [str(e).lower() for e in cm.get("category_examples", []) +
                                              cm.get("department_examples", [])]
    if any(kw in fl for kw in ["品类", "item trading code"]):
        return "Item", "HIGH", "field 名 = Item / 品类"
    if fl == "アイテム":
        return "Item", "HIGH", "field 名 = アイテム"

    # === PRODUCT_CATEGORY — 看 subgroup_codes 的 gender 或 brand_division 含 gender ===
    if fl in ("group", "gender", "department"):
        if any(g in sample_vals_str for g in ["women", "men", "girl", "boy", "baby", "kid", "ladies"]):
            return "PRODUCT_CATEGORY", "HIGH", "value 含 gender keyword"
    if "product team" in fl:
        if any(g in sample_vals_str for g in ["women", "men", "girl", "boy"]):
            return "PRODUCT_CATEGORY", "MEDIUM", "客戶 team 含 gender"

    # === W/K — 從面料 spec 推 ===
    if any(kw in fl for kw in ["body fabric", "main fabric", "main material", "shell fabric", "fabric content"]):
        return "W/K", "MEDIUM", "面料 spec → 推 Knit/Woven"
    if fl == "design type":
        if any(kw in sample_vals_str for kw in ["knit", "woven"]):
            return "W/K", "HIGH", "value 含 Knit/Woven"

    # default
    return "(no map)", "LOW", ""


def main():
    if not SUMMARY.exists():
        print(f"[FAIL] {SUMMARY} not found", file=sys.stderr); return 1
    if not MAPPING_JSON.exists():
        print(f"[FAIL] {MAPPING_JSON} not found", file=sys.stderr); return 1

    summary = json.load(open(SUMMARY, encoding="utf-8"))
    mapping = json.load(open(MAPPING_JSON, encoding="utf-8"))
    print(f"Loaded {len(summary)} clients from PDF universe")
    print(f"Loaded {len(mapping.get('clients', {}))} clients from mapping JSON")

    # Top 22
    top_clients = sorted(summary.keys(), key=lambda c: -summary[c]["n_pdfs"])[:22]
    print(f"Top 22: {top_clients}")

    # Filter to high-cov
    client_fields = {}
    all_fields = set()
    for c in top_clients:
        keys = summary[c]["keys"]
        kept = {k: v for k, v in keys.items() if v["coverage_pct"] >= MIN_COV}
        client_fields[c] = kept
        all_fields.update(kept.keys())
    all_fields = sorted(all_fields)
    print(f"High-cov unique fields: {len(all_fields)}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl required: pip install openpyxl"); return 1

    wb = Workbook()
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")
    GRAY = PatternFill("solid", fgColor="EEEEEE")
    BOLD_GRAY = PatternFill("solid", fgColor="DDDDDD")

    # === Sheet 1: client_field_universe ===
    ws1 = wb.active
    ws1.title = "client_field_universe"
    ws1.append(["client", "n_pdfs"] + all_fields)
    for cell in ws1[1]:
        cell.font = Font(bold=True); cell.fill = BOLD_GRAY
    for c in top_clients:
        n = summary[c]["n_pdfs"]
        row = [c, n]
        for fld in all_fields:
            stat = client_fields[c].get(fld)
            row.append(stat["coverage_pct"] if stat else "")
        ws1.append(row)
    for r in range(2, len(top_clients) + 2):
        for col in range(3, len(all_fields) + 3):
            cell = ws1.cell(row=r, column=col)
            v = cell.value
            if v == "" or v is None: cell.fill = GRAY
            elif isinstance(v, (int, float)):
                if v >= 70: cell.fill = GREEN
                elif v >= 50: cell.fill = YELLOW
                else: cell.fill = RED
    ws1.freeze_panes = "C2"

    # === Sheet 2: field_value_attributes ===
    ws2 = wb.create_sheet("field_value_attributes")
    ws2.append(["client", "field", "n_pdfs_with_field", "coverage_pct",
                "n_unique_values", "top_5_values_with_count", "all_unique_values_sample"])
    for cell in ws2[1]:
        cell.font = Font(bold=True); cell.fill = BOLD_GRAY
    for c in top_clients:
        for fld, stat in sorted(client_fields[c].items(), key=lambda kv: -kv[1]["coverage_pct"]):
            top5 = " | ".join(f"{x['v'][:50]} ({x['count']})" for x in stat["top_5_values"])
            all_uniq = " | ".join(stat["all_unique_values_sample"][:30])
            ws2.append([c, fld, stat["n_pdfs_with_key"], stat["coverage_pct"],
                       stat["n_unique_values"], top5, all_uniq])
    ws2.freeze_panes = "A2"

    # === Sheet 3: client_canonical_xwalk (用 mapping JSON derive) ===
    ws3 = wb.create_sheet("client_canonical_xwalk")
    ws3.append([
        "client", "n_pdfs", "field_name", "coverage_pct", "n_unique_values",
        "top_3_values",
        "→ canonical (M7 列管)", "confidence", "source (mapping JSON 對到的證據)",
        "(IE 部門) final_canonical", "(IE 部門) notes",
    ])
    for cell in ws3[1]:
        cell.font = Font(bold=True); cell.fill = BOLD_GRAY

    n_rows = 0
    by_canon = defaultdict(int)
    by_conf = defaultdict(int)
    for c in top_clients:
        for fld, stat in sorted(client_fields[c].items(), key=lambda kv: -kv[1]["coverage_pct"]):
            canon, conf, src = derive_canonical_from_mapping(c, fld, stat["top_5_values"], mapping)
            top3 = " | ".join(x["v"][:30] for x in stat["top_5_values"][:3])
            ws3.append([
                c, summary[c]["n_pdfs"], fld, stat["coverage_pct"], stat["n_unique_values"],
                top3, canon, conf, src, "", ""
            ])
            n_rows += 1
            by_canon[canon] += 1
            by_conf[conf] += 1

            # Color confidence
            r = n_rows + 1
            cell = ws3.cell(row=r, column=8)
            if conf == "HIGH": cell.fill = GREEN
            elif conf == "MEDIUM": cell.fill = YELLOW
            elif conf == "LOW": cell.fill = RED

            # Color canonical (no map → gray)
            cell2 = ws3.cell(row=r, column=7)
            if canon == "(no map)" or "client 不在" in canon:
                cell2.fill = GRAY

    ws3.freeze_panes = "A2"
    # Column widths
    widths = [22, 8, 28, 10, 10, 70, 25, 10, 70, 25, 30]
    for i, w in enumerate(widths, 1):
        col = chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"
        ws3.column_dimensions[col].width = w

    wb.save(OUT_XLSX)
    print(f"\n✓ Wrote {OUT_XLSX}")
    print(f"  Sheets: client_field_universe / field_value_attributes / client_canonical_xwalk")
    print(f"\n=== Sheet 3 (xwalk) stats ===")
    print(f"  Total rows: {n_rows}")
    print(f"  By canonical:")
    for c, n in sorted(by_canon.items(), key=lambda x: -x[1]):
        print(f"    {c:30} {n:>4}")
    print(f"  By confidence:")
    for c in ("HIGH", "MEDIUM", "LOW"):
        print(f"    {c:8} {by_conf.get(c,0):>4}")


if __name__ == "__main__":
    sys.exit(main() or 0)
